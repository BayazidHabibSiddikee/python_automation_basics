import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python312\\automate resources\\produceSales.xlsx')
sheet = wb.active

update ={'Garlic': 3.07,
         'Celery': 1.19,
         'Lemon':1.27}

for i in range(1,sheet.max_row+1):
    name = sheet.cell(row=i,column = 1).value
    if name in update:
        sheet.cell(row=i,column= 2).value = update[name] #It will replace the value wow simplier than mine

wb.save('ch12.p2.y updated by book.xlsx')
