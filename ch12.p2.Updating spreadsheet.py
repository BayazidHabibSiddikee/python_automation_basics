#! python3
import openpyxl
#wb = openpyxl.Workbook('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python312\\automate resources\\produceSales.xlsx')
#It will create new one -_-
wb = openpyxl.load_workbook('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python312\\automate resources\\produceSales.xlsx')
sheet = wb.active
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

for i in range(1,sheet.max_row+1):
    sheet_new['a'+str(i)] = sheet['a'+str(i)].value
    if sheet['a'+str(i)].value=='Celery':
        sheet_new['b'+ str(i)] = 1.19
    elif sheet['a'+str(i)].value=='Garlic':
        sheet_new['b'+str(i)]= 3.07
    elif sheet['a'+str(i)].value=='':
        sheet_new['b'+str(i)]=1.27
    else:
        sheet_new['b'+str(i)] = sheet['b'+str(i)].value
    sheet_new['c'+str(i)] = sheet['c'+str(i)].value
    sheet_new['d'+str(i)] = sheet['d'+str(i)].value

wb_new.save('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python312\\ch12.p2.Produce_Sales_Updated_1.xlsx')

