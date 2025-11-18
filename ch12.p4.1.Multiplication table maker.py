import openpyxl
wb = openpyxl.Workbook()
#ws = wb.sheetnames[0]
ws=wb.active

try:
    n = int(input("Enter a number: "))
except Exception as e:
    print(f'Try again bruh and {e}')

#Lets create an heading
head = f'Multiplication table of {n}*{n} is :'
ws.merge_cells('a1:h1')
ws['a1']=head

from openpyxl.styles import Font
ws['a1'].font = Font(name='Times New Roman',size = 24,bold = True)




for i in range(1,int(n)+1):
    #ws['a' + str(i)]=i #To create the row why do I forget I have cell coordinates
    ws.cell(row=i+2,column=1).value=i  #left header
    ws.cell(row=2,column =i+1).value=i  #top header
    ws.row_dimensions[i+1].height=25
    
    ws.column_dimensions[chr(96+i)].width=8
    
    for j in range(1,(n)+1):
        exp=f'{i*j}'
        ws.cell(row=i+2,column=j+1).value=int(exp)   #f'{i*j}'   #int(n)
        
    ws.column_dimensions[chr(96+i+1)].width=8
    ws.row_dimensions[i+2].height=25


wb.save('ch12.p4.1.Table.xlsx')
