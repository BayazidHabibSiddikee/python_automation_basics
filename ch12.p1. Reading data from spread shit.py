#! python3
import openpyxl, pprint
#from openpyxl.utils import 

path = 'C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python312\\automate resources\\censuspopdata.xlsx'

print("Opening workbook at: "+path)
wb = openpyxl.load_workbook(path)
sheet = wb[wb.sheetnames[0]]  #<Worksheet "Population by Census Tract">

#Making a dictionary to save this shits
countrydata={}  #countrydata={'state':{'country':{'pop':number,'tracts':number}}}

#making a file to save this
file = open('census2010_result.py','w')

for row in range(2, sheet.max_row+1):
    state = sheet['B' + str(row)].value ##sheet['B2'].value
    country =sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value
    #print(f'{state}: {country} : {pop}')


    #Lets check if the key for the state exists
    countrydata.setdefault(state,{})
    #check if the key for the country exists
    countrydata[state].setdefault(country,{'tracts':0,'pop':0})
    
    #Lets increment the value for tracts and pop
    countrydata[state][country]['tracts']+=1 #how many tracts each country has
    countrydata[state][country]['pop']+=int(pop)
    

#lets save all this data
print("Writing results on file: census2010_result.py")
file.write('data = ' + pprint.pformat(countrydata)+'\n')
file.close()
print("huhhhhhhh done")
    
