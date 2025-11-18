import openpyxl
wb = openpyxl.load_workbook('ch12.p4.1.Table.xlsx')
ws=wb.active
n = int(input("Number of Row:"))
m =int(input('Starting row: '))

for i in range(ws.max_row,m-1,-1):
    for j in range(1,ws.max_column):
        source = ws.cell(row=i,column=j).value
        #if i>=m and i<=(m+n): #if total 1<=i>=12 m = 3, n=4
        ws.cell(row=i+n,column=j).value = source
        ws.cell(row=i,column=j).value=None
        #else:
         #   ws.cell(row=i,column=j).value = source

wb.save('ch12.p5.2.x result.xlsx')
