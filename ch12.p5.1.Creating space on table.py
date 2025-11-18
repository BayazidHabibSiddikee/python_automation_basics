import openpyxl

wb=openpyxl.load_workbook('ch12.p4.1.Table.xlsx')
ws=wb.active
n = int(input("Number of Row:"))
m =int(input('Starting row: '))

'''
for i in range(2,ws.max_row+1+n): #Starting from row 2 to skip merged cell for this one
    if i>=m:
        for k in range(i,n+ws.max_row+1):
            for j in range(1,ws.max_column+1):
                ws.cell(row=k+n,column=j).value =ws.cell(row=k,column=j).value
                ws.cell(row=k,column=j).value =''
    else:
        for j in range(1,ws.max_column+1):
            ws.cell(row=i,column=j).value =ws.cell(row=i,column=j).value
    for j in range(1,ws.max_column+1):
        ws.cell(row=i,column=j).value =ws.cell(row=i,column=j).value
        if i>=m:
            ws.cell(row=i+m,column=j).value =ws.cell(row=i,column=j).value
            for k in range(i,i+m+1):
                ws.cell(row=k+m,column=j).value =ws.cell(row=i,column=j).value
                ws.cell(row=k,column=j).value =''
            #ws.cell(row=i,column=j).value =ws.cell(row=i,column=j).value'''

for i in range(ws.max_row, m - 1, -1):
    for j in range(1, ws.max_column + 1):
        source = ws.cell(row=i, column=j).value
        ws.cell(row=i + n, column=j).value = source
        ws.cell(row=i, column=j).value = None  # Optional: clear old cell if needed
        
wb.save('ch12.p5.1.x result.xlsx')
